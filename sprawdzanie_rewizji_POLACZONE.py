# -*- coding: utf-8 -*-
"""
SPRAWDZANIE REWIZJI - skrypt POŁĄCZONY (1 plik zamiast 2)
=========================================================
Łączy w jeden przebieg:
  FAZA 1 = sprawdzanie_rewizji.py
           (propozycje.xlsx -> wb w pamięci)
           ZEFZEV + split propozycji + nowe_oznaczenie + skrot
           + kopia arkusza 'stary' + wstępny VLOOKUP stare_oznaczenie
  FAZA 2 = dopasowanie_rewizji+REWIZJA.py
           pełny słownik z całego pliku STARE (data_only, wszystkie wiersze)
           -> dokładne stare_oznaczenie + kolumna REWIZJA

Uruchomienie:
  python sprawdzanie_rewizji_POLACZONE.py

JEDYNY plik wyjściowy:
  propozycje_updated_z_dopasowaniem(NNNN).xlsx
    gdzie NNNN = ostatnie 4 cyfry numeru zamówienia z kolumny 'ZLECENIE'
    (np. 45675388 -> 5388). Gdy brak numeru -> propozycje_updated_z_dopasowaniem().xlsx
    (skrypt się NIE zatrzymuje).

Pliki pomocnicze (CSV, plik pośredni propozycje_updated.xlsx) NIE są zostawiane -
plik pośredni jest kasowany na końcu, CSV-ów w ogóle nie tworzymy.

Logika obu faz jest IDENTYCZNA jak w oryginalnych skryptach (z zachowanymi
poprawkami FIX:). Zmienia się tylko: brak plików pobocznych + zmiana nazwy wyniku.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import openpyxl.utils.cell
import os


# Wspólna ścieżka do bazy STARE (używana w obu fazach)
STARE_FILE = r'\\QNAP-ENERGO\Technologia\BAZA\POBIERANIE Z BAZY\SPRAWDZANIE REWIZJI\SPRAWDZANIE_REWIZJI-STARE.xlsx'

# Plik pośredni (kasowany na końcu) i bazowa nazwa wyniku
INTERMEDIATE_FILE = 'propozycje_updated.xlsx'
OUTPUT_BASE = 'propozycje_updated_z_dopasowaniem'  # sufiks (NNNN) dokładany w czasie zmiany nazwy


def is_cell_empty(value):
    """
    Sprawdza czy wartość komórki jest naprawdę pusta.
    """
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == '':
        return True
    return False


def order_suffix_from_zlecenie(ws):
    """
    Zwraca ostatnie 4 cyfry numeru zamówienia z kolumny 'ZLECENIE'
    (np. '45675388' -> '5388'). Gdy kolumny brak lub same puste -> '' .
    Bierze pierwszą niepustą wartość z kolumny ZLECENIE.
    """
    headers = [c.value for c in ws[1]]
    if 'ZLECENIE' not in headers:
        return ''
    col = headers.index('ZLECENIE') + 1
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row=row, column=col).value
        if not is_cell_empty(v):
            digits = ''.join(ch for ch in str(v) if ch.isdigit())
            if len(digits) >= 4:
                return digits[-4:]
            return digits  # mniej niż 4 cyfry - oddaj co jest
    return ''


# ============================================================================
# FAZA 1 -- sprawdzanie_rewizji.py  (zwraca workbook w pamięci, bez zapisu na dysk)
# ============================================================================
def faza1_sprawdzanie():
    # File paths
    propozycje_file = 'propozycje.xlsx'
    stare_file = STARE_FILE

    print("=" * 60)
    print("FAZA 1: SPRAWDZANIE REWIZJI - Przetwarzanie propozycji")
    print("=" * 60)

    # Wczytaj plik Excel używając pandas
    df = pd.read_excel(propozycje_file)

    # Automatycznie znajdź kolumny po nazwach
    try:
        zeinr_col_idx = df.columns.get_loc("Zeinr")
        zef_col_idx = df.columns.get_loc("ZEF")
        zev_col_idx = df.columns.get_loc("ZEV")
    except KeyError as e:
        print(f"BŁĄD: Nie znaleziono kolumny {e}. Sprawdź nazwy kolumn w pliku.")
        exit()

    # Konwertuj indeksy kolumn na litery Excel
    zeinr_col = get_column_letter(zeinr_col_idx + 1)
    zef_col = get_column_letter(zef_col_idx + 1)
    zev_col = get_column_letter(zev_col_idx + 1)

    print(f"\n[1] Znaleziono kolumny podstawowe:")
    print(f"  Zeinr: {zeinr_col} (kolumna {zeinr_col_idx + 1})")
    print(f"  ZEF: {zef_col} (kolumna {zef_col_idx + 1})")
    print(f"  ZEV: {zev_col} (kolumna {zev_col_idx + 1})")

    # Znajdź kolumnę 'propozycja'
    try:
        propozycja_col_idx = df.columns.get_loc("propozycja")
        propozycja_col = get_column_letter(propozycja_col_idx + 1)
        print(f"  propozycja: {propozycja_col} (kolumna {propozycja_col_idx + 1})")
    except KeyError:
        print("OSTRZEŻENIE: Nie znaleziono kolumny 'propozycja'.")
        propozycja_col = None

    # Wczytaj plik używając openpyxl
    wb_propozycje = load_workbook(propozycje_file)
    ws_propozycje = wb_propozycje.active

    # KLUCZOWE: Znajdź ostatnią kolumnę z danymi
    original_max_col = ws_propozycje.max_column
    print(f"\n[2] Oryginalny max_column: {original_max_col} ({get_column_letter(original_max_col)})")

    # ============================================================
    # KROK 1: Dodaj kolumnę ZEFZEV NA KOŃCU
    # ============================================================
    next_col_num = original_max_col + 1
    zefzev_col = get_column_letter(next_col_num)
    ws_propozycje[f'{zefzev_col}1'].value = 'ZEFZEV'
    print(f"\n[3] Utworzono kolumnę ZEFZEV w: {zefzev_col} (kolumna {next_col_num})")

    # Wypełnij ZEFZEV
    # Reguła (jak ręcznie w Excelu):
    #  - ZŁOŻENIE = wiersz z POGRUBIONYM Zeinr. Tam ZEFZEV = ZEF & ZEV
    #    (np. 2+0 -> "20", 1+A -> "1A"); gdy ZEF i ZEV oba puste -> "000" (sentinel).
    #  - pozostałe wiersze (pozycje złożenia) DZIEDZICZĄ wartość złożenia powyżej
    #    (wypełnienie w dół / "fill blanks = wartość nad").
    print("    Wypełnianie ZEFZEV (złożenia=pogrubione + wypełnienie w dół)...")

    def _cell_str(v):
        # liczbę całkowitą zapisz bez '.0'
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v)

    def _zefzev_from(zef, zev):
        zef_e = is_cell_empty(zef)
        zev_e = is_cell_empty(zev)
        if zef_e and zev_e:
            return '000'  # złożenie bez rewizji - sentinel zatrzymujący dziedziczenie
        s = ('' if zef_e else _cell_str(zef)) + ('' if zev_e else _cell_str(zev))
        return s if s != '' else '000'

    zefzev_filled = 0
    zlozenia_cnt = 0
    last_zefzev = None
    for row in range(2, ws_propozycje.max_row + 1):
        zeinr_cell = ws_propozycje[f'{zeinr_col}{row}']
        is_assembly = (not is_cell_empty(zeinr_cell.value)) and bool(zeinr_cell.font and zeinr_cell.font.bold)

        if is_assembly:
            zef = ws_propozycje[f'{zef_col}{row}'].value
            zev = ws_propozycje[f'{zev_col}{row}'].value
            last_zefzev = _zefzev_from(zef, zev)
            ws_propozycje[f'{zefzev_col}{row}'].value = last_zefzev
            zefzev_filled += 1
            zlozenia_cnt += 1
        else:
            # pozycja złożenia - dziedziczy wartość złożenia powyżej
            if last_zefzev is not None:
                ws_propozycje[f'{zefzev_col}{row}'].value = last_zefzev
                zefzev_filled += 1

    print(f"    ✓ Złożeń (pogrubione): {zlozenia_cnt}; wypełniono ZEFZEV w {zefzev_filled} wierszach")

    # ============================================================
    # KROK 2: SPLIT propozycja NA KOŃCU (po ZEFZEV)
    # ============================================================
    split_count = 0
    if propozycja_col:
        next_col_num = ws_propozycje.max_column + 1
        split_start_col_num = next_col_num
        split_start_col = get_column_letter(split_start_col_num)

        print(f"\n[5] SPLIT kolumny propozycja rozpoczyna się od: {split_start_col} (kolumna {split_start_col_num})")
        print(f"    Struktura: grubość_gatunek_rysunek_pozycja_szt_technol_zlecenie_index")

        # Najpierw znajdź maksymalną liczbę części po split'cie
        max_parts = 0
        for row in range(2, ws_propozycje.max_row + 1):
            propozycja = ws_propozycje[f'{propozycja_col}{row}'].value
            if propozycja and not is_cell_empty(propozycja):
                parts = str(propozycja).split('_')
                max_parts = max(max_parts, len(parts))

        print(f"    Maksymalna liczba części po split'cie: {max_parts}")

        # Prawidłowe nazwy nagłówków zgodne ze strukturą
        split_headers = ['grubość', 'gatunek', 'rysunek', 'pozycja', 'szt', 'technol', 'zlecenie', 'index']

        # Utwórz nagłówki dla kolumn split'u
        for i in range(max_parts):
            col_num = split_start_col_num + i
            col_letter = get_column_letter(col_num)
            header_name = split_headers[i] if i < len(split_headers) else f'part{i + 1}'
            ws_propozycje[f'{col_letter}1'].value = header_name
            print(f"    Kolumna {col_letter}: {header_name}")

        # Wykonaj split
        for row in range(2, ws_propozycje.max_row + 1):
            propozycja = ws_propozycje[f'{propozycja_col}{row}'].value
            if propozycja and not is_cell_empty(propozycja):
                parts = str(propozycja).split('_')
                for i, part in enumerate(parts):
                    col_num = split_start_col_num + i
                    col_letter = get_column_letter(col_num)
                    ws_propozycje[f'{col_letter}{row}'].value = part
                split_count += 1

                # Debug: pokaż pierwszy przykład
                if split_count == 1:
                    print(f"\n    Przykład split (wiersz {row}): '{propozycja}'")
                    for i, part in enumerate(parts):
                        header = split_headers[i] if i < len(split_headers) else f'part{i + 1}'
                        print(f"      {header}: {part}")

        print(f"\n    ✓ Wykonano split dla {split_count} wierszy")

        # KLUCZOWE: Zapamiętaj pozycje kolumn rysunek i zlecenie
        # rysunek = 3cia kolumna split (indeks 2)
        # zlecenie = 7ma kolumna split (indeks 6)
        rysunek_col = get_column_letter(split_start_col_num + 2)  # +2 bo rysunek ma indeks 2
        zlecenie_col = get_column_letter(split_start_col_num + 6)  # +6 bo zlecenie ma indeks 6

        print(f"\n    Kolumny dla skrótu:")
        print(f"      rysunek: {rysunek_col} (kolumna {split_start_col_num + 2})")
        print(f"      zlecenie: {zlecenie_col} (kolumna {split_start_col_num + 6})")

    # ============================================================
    # KROK 3: Dodaj nowe_oznaczenie NA KOŃCU
    # ============================================================
    next_col_num = ws_propozycje.max_column + 1
    nowe_oznaczenie_col = get_column_letter(next_col_num)
    ws_propozycje[f'{nowe_oznaczenie_col}1'].value = 'nowe_oznaczenie'
    print(f"\n[6] Utworzono kolumnę nowe_oznaczenie w: {nowe_oznaczenie_col} (kolumna {next_col_num})")

    # Wypełnij nowe_oznaczenie = ZEFZEV (wartość złożenia, wypełniona w dół)
    # tylko tam, gdzie 'propozycja' jest niepusta (pozycje do sprawdzenia rewizji).
    nowe_oznaczenie_filled = 0
    for row in range(2, ws_propozycje.max_row + 1):
        propozycja = ws_propozycje[f'{propozycja_col}{row}'].value if propozycja_col else True

        if not is_cell_empty(propozycja):
            zefzev_val = ws_propozycje[f'{zefzev_col}{row}'].value
            if not is_cell_empty(zefzev_val):
                ws_propozycje[f'{nowe_oznaczenie_col}{row}'].value = zefzev_val
                nowe_oznaczenie_filled += 1

    print(f"    ✓ Wypełniono {nowe_oznaczenie_filled} wierszy")

    # ============================================================
    # KROK 4: Dodaj stare_oznaczenie NA KOŃCU
    # ============================================================
    next_col_num = ws_propozycje.max_column + 1
    stare_oznaczenie_col = get_column_letter(next_col_num)
    ws_propozycje[f'{stare_oznaczenie_col}1'].value = 'stare_oznaczenie'
    print(f"\n[7] Utworzono kolumnę stare_oznaczenie w: {stare_oznaczenie_col} (kolumna {next_col_num})")

    # ============================================================
    # KROK 5: Dodaj skrot NA KOŃCU
    # ============================================================
    next_col_num = ws_propozycje.max_column + 1
    skrot_col = get_column_letter(next_col_num)
    ws_propozycje[f'{skrot_col}1'].value = 'skrot'
    print(f"\n[8] Utworzono kolumnę skrot w: {skrot_col} (kolumna {next_col_num})")

    # Wypełnij skrot: zlecenie + "," + rysunek
    # Przykład: 6154,SL10341218
    skrot_filled = 0
    if propozycja_col:
        print(f"    Wypełnianie skrot (format: zlecenie,rysunek)...")

        for row in range(2, ws_propozycje.max_row + 1):
            propozycja = ws_propozycje[f'{propozycja_col}{row}'].value
            if propozycja and not is_cell_empty(propozycja):
                zlecenie_val = ws_propozycje[f'{zlecenie_col}{row}'].value
                rysunek_val = ws_propozycje[f'{rysunek_col}{row}'].value

                if zlecenie_val and rysunek_val and not is_cell_empty(zlecenie_val) and not is_cell_empty(rysunek_val):
                    # Tworzymy skrot: zlecenie,rysunek
                    ws_propozycje[f'{skrot_col}{row}'].value = str(zlecenie_val) + ',' + str(rysunek_val)
                    skrot_filled += 1

                    # Debug: pokaż pierwsze 5 przykładów
                    if skrot_filled <= 5:
                        print(f"    Wiersz {row}: zlecenie='{zlecenie_val}' + rysunek='{rysunek_val}' = '{str(zlecenie_val)},{str(rysunek_val)}'")

        print(f"    ✓ Wypełniono {skrot_filled} wierszy")

    # ============================================================
    # KROK 6: Kopiowanie arkusza 'stary'
    # ============================================================
    print(f"\n[9] Kopiowanie danych ze starego pliku...")
    if not os.path.exists(stare_file):
        print(f"    BŁĄD: Nie znaleziono pliku '{stare_file}'")
        print(f"    Pomijam kopiowanie arkusza 'stary'")
    else:
        wb_stare = load_workbook(stare_file)
        ws_stare_source = wb_stare.active

        # Usuń arkusz 'stary' jeśli istnieje
        if 'stary' in wb_propozycje.sheetnames:
            wb_propozycje.remove(wb_propozycje['stary'])

        ws_stary = wb_propozycje.create_sheet('stary')

        # Kopiuj kolumny A:F
        copied_rows = 0
        for col in range(1, 7):  # A to F
            col_letter = get_column_letter(col)
            for row in range(1, ws_stare_source.max_row + 1):
                ws_stary[f'{col_letter}{row}'].value = ws_stare_source[f'{col_letter}{row}'].value
                if row == 1:
                    copied_rows = ws_stare_source.max_row - 1

        print(f"    ✓ Skopiowano {copied_rows} wierszy (kolumny A:F)")

    # ============================================================
    # KROK 7: VLOOKUP do stare_oznaczenie
    # ============================================================
    if propozycja_col and 'stary' in wb_propozycje.sheetnames:
        print(f"\n[10] Wypełnianie kolumny stare_oznaczenie (VLOOKUP)...")
        ws_stary = wb_propozycje['stary']

        vlookup_count = 0
        not_found_count = 0

        for row in range(2, ws_propozycje.max_row + 1):
            propozycja = ws_propozycje[f'{propozycja_col}{row}'].value
            if propozycja and not is_cell_empty(propozycja):
                skrot_val = ws_propozycje[f'{skrot_col}{row}'].value
                if skrot_val and not is_cell_empty(skrot_val):
                    # VLOOKUP w arkuszu stary (kolumna C zawiera skrot, kolumna F zawiera stare_oznaczenie)
                    found = False
                    for stary_row in range(2, ws_stary.max_row + 1):
                        if ws_stary[f'C{stary_row}'].value == skrot_val:
                            ws_propozycje[f'{stare_oznaczenie_col}{row}'].value = ws_stary[f'F{stary_row}'].value
                            vlookup_count += 1
                            found = True

                            # Debug: pokaż pierwsze 5 przykładów
                            if vlookup_count <= 5:
                                print(f"    ✓ Wiersz {row}: skrot='{skrot_val}' -> stare='{ws_stary[f'F{stary_row}'].value}'")
                            break

                    if not found:
                        not_found_count += 1
                        if not_found_count <= 3:
                            print(f"    ✗ Wiersz {row}: Nie znaleziono skrótu '{skrot_val}' w arkuszu stary")

        print(f"\n    ✓ Znaleziono {vlookup_count} dopasowań")
        if not_found_count > 0:
            print(f"    ⚠ Nie znaleziono {not_found_count} wpisów w słowniku")

    # ============================================================
    # ZAPIS pliku pośredniego (potrzebny FAZIE 2; kasowany na końcu)
    # ============================================================
    print(f"\n[11] Zapis pliku pośredniego '{INTERMEDIATE_FILE}'...")
    wb_propozycje.save(INTERMEDIATE_FILE)

    print("\n" + "=" * 60)
    print("✓ FAZA 1 UKOŃCZONA POMYŚLNIE")
    print("=" * 60)
    print(f"\n📊 PODSUMOWANIE FAZY 1:")
    print(f"  ZEFZEV: {zefzev_filled} wierszy")
    if propozycja_col:
        print(f"  SPLIT propozycji: {split_count} wierszy")
        print(f"  skrot (zlecenie,rysunek): {skrot_filled} wierszy")
    print(f"  nowe_oznaczenie: {nowe_oznaczenie_filled} wierszy")
    if 'vlookup_count' in locals():
        print(f"  stare_oznaczenie (VLOOKUP): {vlookup_count} dopasowań")
        if 'not_found_count' in locals() and not_found_count > 0:
            print(f"                            {not_found_count} nie znaleziono")
    print(f"   Oryginalne kolumny: {original_max_col} -> po przetworzeniu: {ws_propozycje.max_column}")


# ============================================================================
# FAZA 2 -- dopasowanie_rewizji+REWIZJA.py  (bez plików CSV)
# ============================================================================
def faza2_dopasowanie():
    # File paths
    source_file = INTERMEDIATE_FILE
    stare_file = STARE_FILE

    print("\n\n" + "=" * 60)
    print("FAZA 2: DOPASOWANIE REWIZJI + kolumna REWIZJA")
    print("=" * 60)

    # Sprawdź czy pliki źródłowe istnieją
    if not os.path.exists(source_file):
        print(f"BŁĄD: Nie znaleziono pliku '{source_file}'")
        print("Najpierw uruchom FAZĘ 1 (sprawdzanie).")
        exit()

    if not os.path.exists(stare_file):
        print(f"BŁĄD: Nie znaleziono pliku '{stare_file}'")
        exit()

    # KLUCZOWE: Wczytaj ORYGINALNY plik STARE bezpośrednio Z data_only=True
    print(f"\n=== WCZYTYWANIE ORYGINALNEGO PLIKU STARE ===")
    print(f"Wczytywanie '{stare_file}' z data_only=True...")
    wb_stare_original = load_workbook(stare_file, data_only=True)
    ws_stare_original = wb_stare_original.active

    print(f"  ✓ Wczytano plik STARE (arkusz '{ws_stare_original.title}', max_row={ws_stare_original.max_row})")

    # KROK 1: Tworzenie słownika z ORYGINALNEGO pliku STARE
    # FIX: plik STARE ma >30000 wierszy (np. 33519). Stały cap 30000 ucinał słownik
    # i gubił klucze z końcowych wierszy (np. '3902,SL40034102' w wierszu 31618).
    # Skanujemy WSZYSTKIE wiersze arkusza.
    MAX_ROWS_TO_CHECK = ws_stare_original.max_row

    print(f"\n=== KROK 1: Tworzenie słownika z oryginalnego pliku STARE (kolumny C -> F) ===")
    print(f"  ℹ Przetwarzanie wierszy od 2 do {MAX_ROWS_TO_CHECK}...")

    stary_dict = {}
    empty_keys = 0
    empty_values = 0
    duplicate_keys = 0
    processed_rows = 0
    rows_with_data = 0

    for row in range(2, MAX_ROWS_TO_CHECK + 1):
        processed_rows += 1

        # Progress bar co 2000 wierszy
        if processed_rows % 2000 == 0:
            print(f"  ...przetworzono {processed_rows} wierszy, znaleziono {rows_with_data} z danymi")

        # Czytamy WARTOŚCI z kolumn C i F z ORYGINALNEGO pliku
        key_c = ws_stare_original[f'C{row}'].value
        value_f = ws_stare_original[f'F{row}'].value

        # Jeżeli kolumna C ma dane, zapisujemy
        if not is_cell_empty(key_c):
            rows_with_data += 1

            # Normalizuj klucz
            key_normalized = str(key_c).strip()

            # Sprawdź czy wartość F nie jest pusta
            if is_cell_empty(value_f):
                empty_values += 1

            # Sprawdź duplikaty
            if key_normalized in stary_dict:
                duplicate_keys += 1

            # Dodaj do słownika - ale zachowaj PIERWSZĄ NIEPUSTĄ wartość
            # (jak Excel WYSZUKAJ.PIONOWO = pierwsze trafienie). Ten sam klucz potrafi się
            # powtarzać z pustym 'suma' w późniejszym wierszu - puste NIE może nadpisać wartości.
            if key_normalized not in stary_dict:
                stary_dict[key_normalized] = value_f
            elif is_cell_empty(stary_dict[key_normalized]) and not is_cell_empty(value_f):
                stary_dict[key_normalized] = value_f
        else:
            empty_keys += 1
            # FIX: usunięto przedwczesne przerwanie na 1000 pustych wierszy -
            # ucinało słownik (puste wiersze w środku pliku kończyły skanowanie za wcześnie).

    print(f"\n  ✓ Przetworzono {processed_rows} wierszy, {rows_with_data} z danymi w kolumnie C")
    print(f"  ✓ Słownik: {len(stary_dict)} UNIKALNYCH kluczy (pominięto {empty_keys} pustych, {duplicate_keys} duplikatów, {empty_values} pustych wartości F)")

    # Wczytaj plik pośredni do ZAPISU (pracujemy bezpośrednio na nim)
    print(f"\nWczytywanie '{source_file}' do zapisu...")
    wb_write = load_workbook(source_file)
    ws_write = wb_write.active

    # Znajdź kolumny
    print("\n=== SZUKANIE KOLUMN ===")
    stare_oznaczenie_col = None
    nowe_oznaczenie_col = None
    skrot_col = None

    for col in range(1, ws_write.max_column + 1):
        col_letter = get_column_letter(col)
        header = ws_write[f'{col_letter}1'].value
        if header == 'stare_oznaczenie':
            stare_oznaczenie_col = col_letter
            print(f"  'stare_oznaczenie' w kolumnie: {col_letter}")
        elif header == 'nowe_oznaczenie':
            nowe_oznaczenie_col = col_letter
            print(f"  'nowe_oznaczenie' w kolumnie: {col_letter}")
        elif header == 'skrot':
            skrot_col = col_letter
            print(f"  'skrot' w kolumnie: {skrot_col}")

    if not stare_oznaczenie_col or not skrot_col or not nowe_oznaczenie_col:
        print("BŁĄD: Nie znaleziono wymaganych kolumn!")
        exit()

    # WYPEŁNIANIE kolumny stare_oznaczenie
    print(f"\n=== WYPEŁNIANIE KOLUMNY stare_oznaczenie ===")
    filled_count = 0
    not_found_count = 0
    empty_value_count = 0

    for row in range(2, ws_write.max_row + 1):
        skrot_val = ws_write[f'{skrot_col}{row}'].value

        if not is_cell_empty(skrot_val):
            skrot_normalized = str(skrot_val).strip()

            if skrot_normalized in stary_dict:
                value_f = stary_dict[skrot_normalized]

                if not is_cell_empty(value_f):
                    ws_write[f'{stare_oznaczenie_col}{row}'].value = value_f
                    filled_count += 1

                    if filled_count <= 5:
                        print(f"  ✓ Wiersz {row}: '{skrot_normalized}' -> '{value_f}'")
                else:
                    empty_value_count += 1
            else:
                not_found_count += 1

    print(f"\n  ✓ Wypełniono: {filled_count}   ✗ Nie znaleziono: {not_found_count}   ⚠ Wartość F pusta: {empty_value_count}")

    # DODANIE KOLUMNY REWIZJA
    print(f"\n=== TWORZENIE KOLUMNY REWIZJA ===")

    skrot_col_num = openpyxl.utils.cell.column_index_from_string(skrot_col)
    rewizja_col_num = skrot_col_num + 1
    rewizja_col = get_column_letter(rewizja_col_num)

    ws_write[f'{rewizja_col}1'].value = 'REWIZJA'
    print(f"  ✓ Kolumna REWIZJA: {rewizja_col}")

    for row in range(2, ws_write.max_row + 1):
        formula = f'={nowe_oznaczenie_col}{row}={stare_oznaczenie_col}{row}'
        ws_write[f'{rewizja_col}{row}'].value = formula

    print(f"  ✓ Dodano formuły")

    # ============================================================
    # ZAPIS KOŃCOWY + zmiana nazwy z numerem zamówienia (NNNN)
    # ============================================================
    suffix = order_suffix_from_zlecenie(ws_write)  # ostatnie 4 cyfry z kolumny ZLECENIE (lub '')
    final_file = f'{OUTPUT_BASE}({suffix}).xlsx'

    if suffix:
        print(f"\n=== ZAPIS KOŃCOWY ===")
        print(f"  Numer zamówienia (z kolumny ZLECENIE): ostatnie 4 cyfry = '{suffix}'")
    else:
        print(f"\n=== ZAPIS KOŃCOWY ===")
        print(f"  ⚠ Nie znaleziono numeru zamówienia w kolumnie ZLECENIE - nazwa z pustym ()")

    # Jeżeli plik o docelowej nazwie już istnieje - nadpisz
    if os.path.exists(final_file):
        os.remove(final_file)
    wb_write.save(final_file)
    print(f"✓ Zapisano plik końcowy: '{final_file}'")

    return final_file


# ============================================================================
# URUCHOMIENIE: FAZA 1 -> FAZA 2 -> sprzątanie pliku pośredniego
# ============================================================================
if __name__ == '__main__':
    faza1_sprawdzanie()
    final_file = faza2_dopasowanie()

    # Sprzątanie: usuń plik pośredni (zostaje tylko plik końcowy)
    if os.path.exists(INTERMEDIATE_FILE):
        os.remove(INTERMEDIATE_FILE)

    print("\n\n" + "=" * 60)
    print("✓✓ GOTOWE - obie fazy ukończone")
    print("=" * 60)
    print(f"📄 JEDYNY plik wynikowy: {final_file}")
    print(f"   Porównaj kolumny: nowe_oznaczenie  vs  stare_oznaczenie")
