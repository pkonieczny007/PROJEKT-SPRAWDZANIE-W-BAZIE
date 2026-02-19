import pandas as pd
import re
import os
import shutil
from pathlib import Path
from openpyxl import load_workbook

def utworz_wykaz():
    """
    Funkcja tworzy wykaz.xlsx z wybranego pliku Excel Z ZACHOWANIEM FORMATOWANIA.
    Dodaje kolumnę 'Nazwa' = Zeinr_p_posn
    """
    print("\n=== TWORZENIE WYKAZU ===")
    print("Dostępne pliki Excel w bieżącym folderze:")
    
    # Pokaż dostępne pliki .xlsx i .xlsm (z wyłączeniem wykaz.xlsx i propozycje.xlsx)
    pliki = [f for f in os.listdir('.') if (f.endswith('.xlsx') or f.endswith('.xlsm'))
             and f not in ['wykaz.xlsx', 'propozycje.xlsx']]
    
    if not pliki:
        print("❌ Nie znaleziono żadnych plików Excel w folderze!")
        return False
    
    for i, plik in enumerate(pliki, 1):
        print(f"{i}. {plik}")
    
    # Wybór pliku
    try:
        wybor = int(input("\nWybierz numer pliku (lub 0 aby anulować): "))
        if wybor == 0:
            print("Anulowano.")
            return False
        if wybor < 1 or wybor > len(pliki):
            print("❌ Nieprawidłowy wybór!")
            return False
        
        plik_zrodlowy = pliki[wybor - 1]
    except ValueError:
        print("❌ Nieprawidłowa wartość!")
        return False
    
    # KROK 1: Skopiuj oryginalny plik do wykaz.xlsx (zachowuje formatowanie)
    print(f"\n📂 Kopiuję {plik_zrodlowy} -> wykaz.xlsx z zachowaniem formatowania...")
    try:
        shutil.copy2(plik_zrodlowy, 'wykaz.xlsx')
        print("✅ Utworzono kopię z zachowaniem formatowania")
    except Exception as e:
        print(f"❌ Błąd podczas kopiowania pliku: {e}")
        return False
    
    # KROK 2: Wczytaj dane do pandas tylko do analizy i obliczeń
    print(f"📂 Wczytuję dane do analizy...")
    df = pd.read_excel('wykaz.xlsx', dtype=str)
    
    # Sprawdź czy są kolumny Zeinr i Posn
    if 'Zeinr' not in df.columns or 'Posn' not in df.columns:
        print("❌ Błąd: Plik musi zawierać kolumny 'Zeinr' i 'Posn'!")
        print(f"Znalezione kolumny: {', '.join(df.columns)}")
        os.remove('wykaz.xlsx')  # Usuń niepoprawny plik
        return False

    # Znajdź kolumnę 'zakupy' (bez uwzględniania wielkości liter)
    zakupy_col = None
    for col in df.columns:
        if col.lower() == 'zakupy':
            zakupy_col = col
            break

    if not zakupy_col:
        print("❌ Błąd: Nie znaleziono kolumny 'zakupy'!")
        print(f"Znalezione kolumny: {', '.join(df.columns)}")
        os.remove('wykaz.xlsx')  # Usuń niepoprawny plik
        return False

    # Uzupełnij puste wartości
    df['Zeinr'] = df['Zeinr'].fillna('')
    df['Posn'] = df['Posn'].fillna('')
    df[zakupy_col] = df[zakupy_col].fillna('')

    # KROK 3: Oblicz wartości dla kolumny Nazwa
    df['Nazwa'] = ''
    mask = df[zakupy_col].str.lower().str.contains('blacha', na=False)
    df.loc[mask, 'Nazwa'] = df.loc[mask, 'Zeinr'] + '_p' + df.loc[mask, 'Posn']
    
    # Dodaj pustą kolumnę propozycja
    df['propozycja'] = ''
    
    # KROK 4: Zapisz nowe kolumny do pliku z zachowaniem formatowania używając openpyxl
    print("💾 Zapisuję nowe kolumny z zachowaniem formatowania...")
    try:
        wb = load_workbook('wykaz.xlsx')
        ws = wb.active
        
        # Pobierz nagłówki
        header_row = [cell.value for cell in ws[1]]
        
        # Dodaj kolumnę 'Nazwa' jeśli nie istnieje
        if 'Nazwa' not in header_row:
            nazwa_col_idx = len(header_row) + 1
            ws.cell(row=1, column=nazwa_col_idx, value='Nazwa')
        else:
            nazwa_col_idx = header_row.index('Nazwa') + 1
        
        # Dodaj kolumnę 'propozycja' jeśli nie istnieje
        header_row = [cell.value for cell in ws[1]]  # Odśwież nagłówki
        if 'propozycja' not in header_row:
            prop_col_idx = len(header_row) + 1
            ws.cell(row=1, column=prop_col_idx, value='propozycja')
        else:
            prop_col_idx = header_row.index('propozycja') + 1
        
        # Wpisz wartości z DataFrame
        for idx, (nazwa_val, prop_val) in enumerate(zip(df['Nazwa'], df['propozycja']), start=2):
            ws.cell(row=idx, column=nazwa_col_idx, value=nazwa_val)
            ws.cell(row=idx, column=prop_col_idx, value=prop_val)
        
        # Zapisz plik
        wb.save('wykaz.xlsx')
        wb.close()
        print("✅ Zapisano z zachowaniem formatowania!")
        
    except Exception as e:
        print(f"⚠️  Błąd podczas zapisu z openpyxl: {e}")
        print("   Zapisuję standardowo przez pandas (bez formatowania)...")
        df.to_excel('wykaz.xlsx', index=False)
    
    print(f"✅ Utworzono wykaz.xlsx z {len(df)} wierszami")
    print(f"   Kolumny: {', '.join(df.columns)}")
    
    # Policz ile pozycji z 'blacha' znaleziono
    ile_blach = mask.sum()
    print(f"   Znaleziono {ile_blach} pozycji z 'blacha' w kolumnie '{zakupy_col}'")
    
    return True


def uruchom_propozycje():
    """
    Funkcja uruchamia dopasowanie propozycji z pliku elementy1.xlsx
    Tworzy propozycje.xlsx jako kopię wykaz.xlsx (z zachowaniem formatowania) i tam dodaje kolumny
    """
    print("\n=== DOPASOWANIE PROPOZYCJI ===")
    
    # Sprawdź czy istnieje wykaz.xlsx
    if not os.path.exists('wykaz.xlsx'):
        print("❌ Błąd: Nie znaleziono pliku wykaz.xlsx!")
        return False
    
    # Sprawdź czy istnieje elementy1.xlsx
    if not os.path.exists('elementy1.xlsx'):
        print("❌ Nie znaleziono pliku elementy1.xlsx w folderze lokalnym!")
        print("   Pobieram najnowszą wersję z bazy danych...")

        # Ścieżka do pliku źródłowego
        sciezka_zrodlowa = r"\\QNAP-ENERGO\Technologia\BAZA\POBIERANIE Z BAZY\elementy1.xlsx"

        if not os.path.exists(sciezka_zrodlowa):
            print("❌ Błąd: Nie można znaleźć pliku źródłowego w bazie danych!")
            return False

        try:
            # Skopiuj plik
            shutil.copy2(sciezka_zrodlowa, 'elementy1.xlsx')
            print("✅ Pobrano plik elementy1.xlsx z bazy danych")

            # Wczytaj datę utworzenia z komórki B2
            df_data = pd.read_excel('elementy1.xlsx', header=None)
            if df_data.shape[0] > 1 and df_data.shape[1] > 1:
                data_utworzenia = df_data.iloc[1, 1]  # B2
                print(f"📅 Data utworzenia bazy: {data_utworzenia}")
            else:
                print("⚠️  Nie udało się odczytać daty utworzenia z pliku")

        except Exception as e:
            print(f"❌ Błąd podczas pobierania pliku: {e}")
            return False
    
    # KROK 1: Skopiuj wykaz.xlsx do propozycje.xlsx (zachowuje formatowanie)
    print("📋 Tworzę kopię wykaz.xlsx -> propozycje.xlsx z zachowaniem formatowania...")
    try:
        shutil.copy2('wykaz.xlsx', 'propozycje.xlsx')
        print("✅ Utworzono kopię z zachowaniem formatowania")
    except Exception as e:
        print(f"❌ Błąd podczas kopiowania pliku: {e}")
        return False
    
    # KROK 2: Wczytaj dane do pandas dla logiki dopasowania
    print("📂 Wczytuję dane...")
    wykaz_df = pd.read_excel("propozycje.xlsx", dtype=str)
    elementy_df = pd.read_excel("elementy1.xlsx", dtype=str)
    
    # Sprawdź czy są wymagane kolumny
    if 'Nazwa' not in wykaz_df.columns:
        print("❌ Błąd: wykaz.xlsx nie zawiera kolumny 'Nazwa'!")
        return False
    
    if 'Referencja' not in elementy_df.columns:
        print("❌ Błąd: elementy1.xlsx nie zawiera kolumny 'Referencja'!")
        return False
    
    # KROK 3: Przygotowanie kolumny wynikowej
    if 'propozycja' not in wykaz_df.columns:
        wykaz_df['propozycja'] = ""
    
    # Uzupełnij puste pola
    elementy_df['Referencja'] = elementy_df['Referencja'].fillna('')
    
    print(f"🔍 Rozpoczynam dopasowanie dla {len(wykaz_df)} pozycji...")
    dopasowane = 0
    
    # KROK 4: Iteracja po wykazie i dopasowanie
    for idx, row in wykaz_df.iterrows():
        nazwa = row.get('Nazwa')
        if not isinstance(nazwa, str) or nazwa.strip() == "":
            continue
        
        nazwa = nazwa.strip()
        
        # Tworzymy regex wzorzec
        regex = re.escape(nazwa) + r'(?!\d)'
        
        # Szukamy dopasowań
        dopasowania = elementy_df[elementy_df['Referencja'].str.contains(regex, na=False, regex=True)]
        
        if not dopasowania.empty:
            match = dopasowania.iloc[0]
            wykaz_df.at[idx, 'propozycja'] = match.get('Referencja1', "")
            dopasowane += 1
    
    # KROK 5: Zapisz dane do istniejącego pliku z zachowaniem formatowania używając openpyxl
    print("💾 Zapisuję wyniki z zachowaniem formatowania...")
    try:
        wb = load_workbook('propozycje.xlsx')
        ws = wb.active
        
        # Znajdź indeks kolumny 'propozycja'
        header_row = [cell.value for cell in ws[1]]
        
        if 'propozycja' in header_row:
            prop_col_idx = header_row.index('propozycja') + 1
        else:
            # Dodaj nową kolumnę na końcu
            prop_col_idx = len(header_row) + 1
            ws.cell(row=1, column=prop_col_idx, value='propozycja')
        
        # Wpisz wartości z DataFrame
        for idx, value in enumerate(wykaz_df['propozycja'], start=2):
            ws.cell(row=idx, column=prop_col_idx, value=value)
        
        # Zapisz plik
        wb.save('propozycje.xlsx')
        wb.close()
        print("✅ Zapisano z zachowaniem formatowania!")
        
    except Exception as e:
        print(f"⚠️  Błąd podczas zapisu z openpyxl: {e}")
        print("   Zapisuję standardowo przez pandas...")
        wykaz_df.to_excel("propozycje.xlsx", index=False)
    
    print(f"✅ Dopasowanie zakończone!")
    print(f"   Dopasowano: {dopasowane}/{len(wykaz_df)} pozycji")
    print(f"   Wyniki zapisano w: propozycje.xlsx")
    return True


def main():
    """
    Główna funkcja - sprawdza czy wykaz.xlsx istnieje i decyduje co zrobić
    """
    print("=" * 50)
    print("  SKRYPT WYKAZ + PROPOZYCJE")
    print("=" * 50)
    
    if os.path.exists('wykaz.xlsx'):
        print("\n📋 Znaleziono plik wykaz.xlsx")
        print("   Uruchamiam dopasowanie propozycji...\n")
        uruchom_propozycje()
    else:
        print("\n📋 Nie znaleziono pliku wykaz.xlsx")
        print("   Tworzę nowy wykaz...\n")
        if utworz_wykaz():
            print("\n" + "=" * 50)
            print("ℹ️  NASTĘPNY KROK:")
            print("   1. Upewnij się, że masz plik 'elementy1.xlsx'")
            print("   2. Uruchom skrypt ponownie, aby dopasować propozycje")
            print("=" * 50)


if __name__ == "__main__":
    main()
