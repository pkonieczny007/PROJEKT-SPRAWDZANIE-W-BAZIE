from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import openpyxl.utils.cell
import shutil
import os
import csv


def is_cell_empty(value):
    """
    Sprawdza czy wartość komórki jest naprawdę pusta.
    """
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == '':
        return True
    return False


# File paths
source_file = 'propozycje_updated.xlsx'
stare_file = r'\\QNAP-ENERGO\Technologia\BAZA\POBIERANIE Z BAZY\SPRAWDZANIE REWIZJI\SPRAWDZANIE_REWIZJI-STARE.xlsx'
output_file = 'propozycje_updated_z_dopasowaniem.xlsx'
dictionary_export_file = 'slownik_stary_C_F.csv'
all_rows_export_file = 'wszystkie_wiersze_stary_C_F.csv'


# Sprawdź czy pliki źródłowe istnieją
if not os.path.exists(source_file):
    print(f"BŁĄD: Nie znaleziono pliku '{source_file}'")
    print("Najpierw uruchom skrypt 'sprawdzanie_rewizji.py'")
    exit()

if not os.path.exists(stare_file):
    print(f"BŁĄD: Nie znaleziono pliku '{stare_file}'")
    exit()


# KLUCZOWE: Wczytaj ORYGINALNY plik STARE bezpośrednio Z data_only=True
print(f"\n=== WCZYTYWANIE ORYGINALNEGO PLIKU STARE ===")
print(f"Wczytywanie '{stare_file}' z data_only=True...")
wb_stare_original = load_workbook(stare_file, data_only=True)
ws_stare_original = wb_stare_original.active

print(f"  ✓ Wczytano plik STARE")
print(f"  ℹ Nazwa aktywnego arkusza: {ws_stare_original.title}")
print(f"  ℹ max_row = {ws_stare_original.max_row}")
print(f"  ℹ max_column = {ws_stare_original.max_column}")

# Sprawdź nagłówki
print(f"\n  ℹ Nagłówki w pierwszym wierszu:")
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    header = ws_stare_original[f'{col}1'].value
    print(f"    Kolumna {col}: '{header}'")


# KROK 1: Tworzenie słownika z ORYGINALNEGO pliku STARE
# FIX: plik STARE ma >30000 wierszy (np. 33519). Stały cap 30000 ucinał słownik
# i gubił klucze z końcowych wierszy (np. '3902,SL40034102' w wierszu 31618).
# Skanujemy WSZYSTKIE wiersze arkusza.
MAX_ROWS_TO_CHECK = ws_stare_original.max_row

print(f"\n=== KROK 1: Tworzenie słownika z oryginalnego pliku STARE (kolumny C -> F) ===")
print(f"  ℹ Przetwarzanie wierszy od 2 do {MAX_ROWS_TO_CHECK}...")

stary_dict = {}
all_rows_data = []
empty_keys = 0
empty_values = 0
duplicate_keys = 0
processed_rows = 0
rows_with_data = 0


for row in range(2, MAX_ROWS_TO_CHECK + 1):
    processed_rows += 1
    
    # Progress bar co 2000 wierszy
    if processed_rows % 2000 == 0:
        print(f"  ...przetworzono {processed_rows} wierszy, znaleziono {rows_with_data} z danymi")
    
    # Czytamy WARTOŚCI z kolumn C i F z ORYGINALNEGO pliku
    key_c = ws_stare_original[f'C{row}'].value
    value_f = ws_stare_original[f'F{row}'].value
    
    # Jeżeli kolumna C ma dane, zapisujemy
    if not is_cell_empty(key_c):
        rows_with_data += 1
        
        # Normalizuj klucz
        key_normalized = str(key_c).strip()
        
        # Dodaj do listy WSZYSTKICH wierszy
        all_rows_data.append({
            'wiersz': row,
            'klucz_c': key_normalized,
            'wartosc_f': value_f if not is_cell_empty(value_f) else "PUSTA"
        })
        
        # Sprawdź czy wartość F nie jest pusta
        if is_cell_empty(value_f):
            empty_values += 1
            if empty_values <= 5:
                print(f"  ⚠ Wiersz {row}: Klucz '{key_normalized}' ma pustą wartość F")
        
        # Sprawdź duplikaty
        if key_normalized in stary_dict:
            duplicate_keys += 1
        
        # Dodaj do słownika - ale zachowaj PIERWSZĄ NIEPUSTĄ wartość
        # (jak Excel WYSZUKAJ.PIONOWO = pierwsze trafienie). Ten sam klucz potrafi się
        # powtarzać z pustym 'suma' w późniejszym wierszu - puste NIE może nadpisać wartości.
        if key_normalized not in stary_dict:
            stary_dict[key_normalized] = value_f
        elif is_cell_empty(stary_dict[key_normalized]) and not is_cell_empty(value_f):
            stary_dict[key_normalized] = value_f
    else:
        empty_keys += 1
        # FIX: usunięto przedwczesne przerwanie na 1000 pustych wierszy -
        # ucinało słownik (puste wiersze w środku pliku kończyły skanowanie za wcześnie).


print(f"\n  ✓ Przetworzono {processed_rows} wierszy")
print(f"  ✓ Znaleziono {rows_with_data} wierszy z danymi w kolumnie C")
print(f"  ✓ Utworzono słownik z {len(stary_dict)} UNIKALNYMI kluczami")
print(f"  ℹ Pominięto {empty_keys} pustych kluczy")
print(f"  ⚠ Znaleziono {empty_values} pustych wartości (kolumna F)")
print(f"  ℹ Znaleziono {duplicate_keys} duplikatów kluczy")


# EKSPORT WSZYSTKICH WIERSZY
print(f"\n=== EKSPORT WSZYSTKICH WIERSZY DO CSV: {all_rows_export_file} ===")
with open(all_rows_export_file, 'w', newline='', encoding='utf-8') as csvfile:
    writer = csv.writer(csvfile, delimiter=';')
    writer.writerow(['Wiersz_arkusza', 'Klucz_C', 'Wartość_F'])
    
    for row_data in all_rows_data:
        writer.writerow([row_data['wiersz'], row_data['klucz_c'], row_data['wartosc_f']])

print(f"  ✓ Zapisano {len(all_rows_data)} WSZYSTKICH wierszy z danymi")


# EKSPORT SŁOWNIKA
print(f"\n=== EKSPORT SŁOWNIKA DO CSV: {dictionary_export_file} ===")
with open(dictionary_export_file, 'w', newline='', encoding='utf-8') as csvfile:
    writer = csv.writer(csvfile, delimiter=';')
    writer.writerow(['Klucz_C', 'Wartość_F', 'Czy_pusta'])
    
    for key in sorted(stary_dict.keys()):
        is_empty = "PUSTA" if is_cell_empty(stary_dict[key]) else "OK"
        writer.writerow([key, stary_dict[key], is_empty])

print(f"  ✓ Zapisano {len(stary_dict)} UNIKATOWYCH wpisów")


# Przykładowe wpisy
print(f"\n  ℹ Przykładowe wpisy ze słownika (pierwsze 10):")
for i, (k, v) in enumerate(list(stary_dict.items())[:10]):
    status = "✓" if not is_cell_empty(v) else "✗ PUSTA"
    print(f"    [{i+1}] '{k}' -> '{v}' {status}")

# Test konkretnego klucza
test_key = "5529,SL10115513"
if test_key in stary_dict:
    print(f"\n  ✓ TEST: Klucz '{test_key}' ISTNIEJE -> '{stary_dict[test_key]}'")
else:
    print(f"\n  ✗ TEST: Klucz '{test_key}' NIE ISTNIEJE")


# Kopiuj plik propozycje
print(f"\n=== KOPIOWANIE PLIKU PROPOZYCJE ===")
print(f"Kopiowanie '{source_file}' -> '{output_file}'...")
shutil.copy(source_file, output_file)


# Wczytaj skopiowany plik do ZAPISU
print(f"\nWczytywanie '{output_file}' do zapisu...")
wb_write = load_workbook(output_file)
ws_write = wb_write.active


# Znajdź kolumny
print("\n=== SZUKANIE KOLUMN ===")
stare_oznaczenie_col = None
nowe_oznaczenie_col = None
skrot_col = None

for col in range(1, ws_write.max_column + 1):
    col_letter = get_column_letter(col)
    header = ws_write[f'{col_letter}1'].value
    if header == 'stare_oznaczenie':
        stare_oznaczenie_col = col_letter
        print(f"  'stare_oznaczenie' w kolumnie: {col_letter}")
    elif header == 'nowe_oznaczenie':
        nowe_oznaczenie_col = col_letter
        print(f"  'nowe_oznaczenie' w kolumnie: {col_letter}")
    elif header == 'skrot':
        skrot_col = col_letter
        print(f"  'skrot' w kolumnie: {skrot_col}")

if not stare_oznaczenie_col or not skrot_col or not nowe_oznaczenie_col:
    print("BŁĄD: Nie znaleziono wymaganych kolumn!")
    exit()


# EKSPORT WARTOŚCI SKROT
skrot_export_file = 'wartosci_skrot_do_sprawdzenia.csv'
print(f"\n=== EKSPORT WARTOŚCI SKROT: {skrot_export_file} ===")
with open(skrot_export_file, 'w', newline='', encoding='utf-8') as csvfile:
    writer = csv.writer(csvfile, delimiter=';')
    writer.writerow(['Wiersz', 'Wartość_skrot', 'Istnieje_w_słowniku', 'Wartość_F', 'Status'])
    
    for row in range(2, ws_write.max_row + 1):
        skrot_val = ws_write[f'{skrot_col}{row}'].value
        if not is_cell_empty(skrot_val):
            skrot_normalized = str(skrot_val).strip()
            exists = "TAK" if skrot_normalized in stary_dict else "NIE"
            value_f = stary_dict.get(skrot_normalized, "BRAK")
            status = "OK" if exists == "TAK" and not is_cell_empty(value_f) else "PROBLEM"
            writer.writerow([row, skrot_normalized, exists, value_f, status])

print(f"  ✓ Zapisano")


# WYPEŁNIANIE kolumny stare_oznaczenie
print(f"\n=== WYPEŁNIANIE KOLUMNY stare_oznaczenie ===")
filled_count = 0
not_found_count = 0
empty_value_count = 0

for row in range(2, ws_write.max_row + 1):
    skrot_val = ws_write[f'{skrot_col}{row}'].value
    
    if not is_cell_empty(skrot_val):
        skrot_normalized = str(skrot_val).strip()
        
        if skrot_normalized in stary_dict:
            value_f = stary_dict[skrot_normalized]
            
            if not is_cell_empty(value_f):
                ws_write[f'{stare_oznaczenie_col}{row}'].value = value_f
                filled_count += 1
                
                if filled_count <= 5:
                    print(f"  ✓ Wiersz {row}: '{skrot_normalized}' -> '{value_f}'")
            else:
                empty_value_count += 1
        else:
            not_found_count += 1

print(f"\n  ✓ Wypełniono: {filled_count}")
print(f"  ✗ Nie znaleziono: {not_found_count}")
print(f"  ⚠ Wartość F pusta: {empty_value_count}")


# DODANIE KOLUMNY REWIZJA
print(f"\n=== TWORZENIE KOLUMNY REWIZJA ===")

skrot_col_num = openpyxl.utils.cell.column_index_from_string(skrot_col)
rewizja_col_num = skrot_col_num + 1
rewizja_col = get_column_letter(rewizja_col_num)

ws_write[f'{rewizja_col}1'].value = 'REWIZJA'
print(f"  ✓ Kolumna REWIZJA: {rewizja_col}")

for row in range(2, ws_write.max_row + 1):
    formula = f'={nowe_oznaczenie_col}{row}={stare_oznaczenie_col}{row}'
    ws_write[f'{rewizja_col}{row}'].value = formula

print(f"  ✓ Dodano formuły")


# ZAPISYWANIE
print(f"\n=== ZAPISYWANIE ===")
wb_write.save(output_file)
print(f"✓ Zapisano '{output_file}'")

print(f"\n📄 Pliki:")
print(f"  1. {all_rows_export_file} - {len(all_rows_data)} wierszy")
print(f"  2. {dictionary_export_file} - {len(stary_dict)} kluczy")
print(f"  3. {skrot_export_file} - analiza")
print(f"  4. {output_file} - wynik")
