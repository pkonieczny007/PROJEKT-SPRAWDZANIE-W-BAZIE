import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import openpyxl.utils.cell
import os

def is_cell_empty(value):
    """
    Sprawdza czy wartość komórki jest naprawdę pusta.
    """
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == '':
        return True
    return False

# File paths
propozycje_file = 'propozycje.xlsx'
stare_file = r'\\QNAP-ENERGO\Technologia\BAZA\POBIERANIE Z BAZY\SPRAWDZANIE REWIZJI\SPRAWDZANIE_REWIZJI-STARE.xlsx'

print("="*60)
print("SPRAWDZANIE REWIZJI - Przetwarzanie propozycji")
print("="*60)

# Wczytaj plik Excel używając pandas
df = pd.read_excel(propozycje_file)

# Automatycznie znajdź kolumny po nazwach
try:
    zeinr_col_idx = df.columns.get_loc("Zeinr")
    zef_col_idx = df.columns.get_loc("ZEF")
    zev_col_idx = df.columns.get_loc("ZEV")
except KeyError as e:
    print(f"BŁĄD: Nie znaleziono kolumny {e}. Sprawdź nazwy kolumn w pliku.")
    exit()

# Konwertuj indeksy kolumn na litery Excel
zeinr_col = get_column_letter(zeinr_col_idx + 1)
zef_col = get_column_letter(zef_col_idx + 1)
zev_col = get_column_letter(zev_col_idx + 1)

print(f"\n[1] Znaleziono kolumny podstawowe:")
print(f"  Zeinr: {zeinr_col} (kolumna {zeinr_col_idx + 1})")
print(f"  ZEF: {zef_col} (kolumna {zef_col_idx + 1})")
print(f"  ZEV: {zev_col} (kolumna {zev_col_idx + 1})")

# Znajdź kolumnę 'propozycja'
try:
    propozycja_col_idx = df.columns.get_loc("propozycja")
    propozycja_col = get_column_letter(propozycja_col_idx + 1)
    print(f"  propozycja: {propozycja_col} (kolumna {propozycja_col_idx + 1})")
except KeyError:
    print("OSTRZEŻENIE: Nie znaleziono kolumny 'propozycja'.")
    propozycja_col = None

# Wczytaj plik używając openpyxl
wb_propozycje = load_workbook(propozycje_file)
ws_propozycje = wb_propozycje.active

# KLUCZOWE: Znajdź ostatnią kolumnę z danymi
original_max_col = ws_propozycje.max_column
print(f"\n[2] Oryginalny max_column: {original_max_col} ({get_column_letter(original_max_col)})")

# ============================================================
# KROK 1: Dodaj kolumnę ZEFZEV NA KOŃCU
# ============================================================
next_col_num = original_max_col + 1
zefzev_col = get_column_letter(next_col_num)
ws_propozycje[f'{zefzev_col}1'].value = 'ZEFZEV'
print(f"\n[3] Utworzono kolumnę ZEFZEV w: {zefzev_col} (kolumna {next_col_num})")

# Wypełnij ZEFZEV
print("    Wypełnianie ZEFZEV...")
zefzev_filled = 0
for row in range(2, ws_propozycje.max_row + 1):
    zef = ws_propozycje[f'{zef_col}{row}'].value
    zev = ws_propozycje[f'{zev_col}{row}'].value
    
    if not is_cell_empty(zef) and not is_cell_empty(zev):
        ws_propozycje[f'{zefzev_col}{row}'].value = str(zef) + str(zev)
        zefzev_filled += 1

print(f"    ✓ Wypełniono {zefzev_filled} wierszy")

# Tworzenie słownika Zeinr -> ZEFZEV
print("\n[4] Tworzenie słownika Zeinr -> ZEFZEV...")
zeinr_to_zefzev = {}
for row in range(2, ws_propozycje.max_row + 1):
    zeinr = ws_propozycje[f'{zeinr_col}{row}'].value
    zefzev = ws_propozycje[f'{zefzev_col}{row}'].value
    
    if not is_cell_empty(zeinr) and not is_cell_empty(zefzev):
        zeinr_to_zefzev[zeinr] = zefzev

print(f"    ✓ Utworzono słownik z {len(zeinr_to_zefzev)} wpisami")

# ============================================================
# KROK 2: SPLIT propozycja NA KOŃCU (po ZEFZEV)
# ============================================================
if propozycja_col:
    next_col_num = ws_propozycje.max_column + 1
    split_start_col_num = next_col_num
    split_start_col = get_column_letter(split_start_col_num)
    
    print(f"\n[5] SPLIT kolumny propozycja rozpoczyna się od: {split_start_col} (kolumna {split_start_col_num})")
    print(f"    Struktura: grubość_gatunek_rysunek_pozycja_szt_technol_zlecenie_index")
    
    # Najpierw znajdź maksymalną liczbę części po split'cie
    max_parts = 0
    for row in range(2, ws_propozycje.max_row + 1):
        propozycja = ws_propozycje[f'{propozycja_col}{row}'].value
        if propozycja and not is_cell_empty(propozycja):
            parts = str(propozycja).split('_')
            max_parts = max(max_parts, len(parts))
    
    print(f"    Maksymalna liczba części po split'cie: {max_parts}")
    
    # Prawidłowe nazwy nagłówków zgodne ze strukturą
    split_headers = ['grubość', 'gatunek', 'rysunek', 'pozycja', 'szt', 'technol', 'zlecenie', 'index']
    
    # Utwórz nagłówki dla kolumn split'u
    for i in range(max_parts):
        col_num = split_start_col_num + i
        col_letter = get_column_letter(col_num)
        header_name = split_headers[i] if i < len(split_headers) else f'part{i+1}'
        ws_propozycje[f'{col_letter}1'].value = header_name
        print(f"    Kolumna {col_letter}: {header_name}")
    
    # Wykonaj split
    split_count = 0
    for row in range(2, ws_propozycje.max_row + 1):
        propozycja = ws_propozycje[f'{propozycja_col}{row}'].value
        if propozycja and not is_cell_empty(propozycja):
            parts = str(propozycja).split('_')
            for i, part in enumerate(parts):
                col_num = split_start_col_num + i
                col_letter = get_column_letter(col_num)
                ws_propozycje[f'{col_letter}{row}'].value = part
            split_count += 1
            
            # Debug: pokaż pierwszy przykład
            if split_count == 1:
                print(f"\n    Przykład split (wiersz {row}): '{propozycja}'")
                for i, part in enumerate(parts):
                    header = split_headers[i] if i < len(split_headers) else f'part{i+1}'
                    print(f"      {header}: {part}")
    
    print(f"\n    ✓ Wykonano split dla {split_count} wierszy")
    
    # KLUCZOWE: Zapamiętaj pozycje kolumn rysunek i zlecenie
    # rysunek = 3cia kolumna split (indeks 2)
    # zlecenie = 7ma kolumna split (indeks 6)
    rysunek_col = get_column_letter(split_start_col_num + 2)  # +2 bo rysunek ma indeks 2
    zlecenie_col = get_column_letter(split_start_col_num + 6)  # +6 bo zlecenie ma indeks 6
    
    print(f"\n    Kolumny dla skrótu:")
    print(f"      rysunek: {rysunek_col} (kolumna {split_start_col_num + 2})")
    print(f"      zlecenie: {zlecenie_col} (kolumna {split_start_col_num + 6})")

# ============================================================
# KROK 3: Dodaj nowe_oznaczenie NA KOŃCU
# ============================================================
next_col_num = ws_propozycje.max_column + 1
nowe_oznaczenie_col = get_column_letter(next_col_num)
ws_propozycje[f'{nowe_oznaczenie_col}1'].value = 'nowe_oznaczenie'
print(f"\n[6] Utworzono kolumnę nowe_oznaczenie w: {nowe_oznaczenie_col} (kolumna {next_col_num})")

# Wypełnij nowe_oznaczenie
nowe_oznaczenie_filled = 0
for row in range(2, ws_propozycje.max_row + 1):
    zeinr = ws_propozycje[f'{zeinr_col}{row}'].value
    propozycja = ws_propozycje[f'{propozycja_col}{row}'].value if propozycja_col else True
    
    if not is_cell_empty(propozycja):
        if zeinr in zeinr_to_zefzev:
            ws_propozycje[f'{nowe_oznaczenie_col}{row}'].value = zeinr_to_zefzev[zeinr]
            nowe_oznaczenie_filled += 1

print(f"    ✓ Wypełniono {nowe_oznaczenie_filled} wierszy")

# ============================================================
# KROK 4: Dodaj stare_oznaczenie NA KOŃCU
# ============================================================
next_col_num = ws_propozycje.max_column + 1
stare_oznaczenie_col = get_column_letter(next_col_num)
ws_propozycje[f'{stare_oznaczenie_col}1'].value = 'stare_oznaczenie'
print(f"\n[7] Utworzono kolumnę stare_oznaczenie w: {stare_oznaczenie_col} (kolumna {next_col_num})")

# ============================================================
# KROK 5: Dodaj skrot NA KOŃCU
# ============================================================
next_col_num = ws_propozycje.max_column + 1
skrot_col = get_column_letter(next_col_num)
ws_propozycje[f'{skrot_col}1'].value = 'skrot'
print(f"\n[8] Utworzono kolumnę skrot w: {skrot_col} (kolumna {next_col_num})")

# Wypełnij skrot: zlecenie + "," + rysunek
# Przykład: 6154,SL10341218
if propozycja_col:
    skrot_filled = 0
    print(f"    Wypełnianie skrot (format: zlecenie,rysunek)...")
    
    for row in range(2, ws_propozycje.max_row + 1):
        propozycja = ws_propozycje[f'{propozycja_col}{row}'].value
        if propozycja and not is_cell_empty(propozycja):
            zlecenie_val = ws_propozycje[f'{zlecenie_col}{row}'].value
            rysunek_val = ws_propozycje[f'{rysunek_col}{row}'].value
            
            if zlecenie_val and rysunek_val and not is_cell_empty(zlecenie_val) and not is_cell_empty(rysunek_val):
                # Tworzymy skrot: zlecenie,rysunek
                ws_propozycje[f'{skrot_col}{row}'].value = str(zlecenie_val) + ',' + str(rysunek_val)
                skrot_filled += 1
                
                # Debug: pokaż pierwsze 5 przykładów
                if skrot_filled <= 5:
                    print(f"    Wiersz {row}: zlecenie='{zlecenie_val}' + rysunek='{rysunek_val}' = '{str(zlecenie_val)},{str(rysunek_val)}'")
    
    print(f"    ✓ Wypełniono {skrot_filled} wierszy")

# ============================================================
# KROK 6: Kopiowanie arkusza 'stary'
# ============================================================
print(f"\n[9] Kopiowanie danych ze starego pliku...")
if not os.path.exists(stare_file):
    print(f"    BŁĄD: Nie znaleziono pliku '{stare_file}'")
    print(f"    Pomijam kopiowanie arkusza 'stary'")
else:
    wb_stare = load_workbook(stare_file)
    ws_stare_source = wb_stare.active
    
    # Usuń arkusz 'stary' jeśli istnieje
    if 'stary' in wb_propozycje.sheetnames:
        wb_propozycje.remove(wb_propozycje['stary'])
    
    ws_stary = wb_propozycje.create_sheet('stary')
    
    # Kopiuj kolumny A:F
    copied_rows = 0
    for col in range(1, 7):  # A to F
        col_letter = get_column_letter(col)
        for row in range(1, ws_stare_source.max_row + 1):
            ws_stary[f'{col_letter}{row}'].value = ws_stare_source[f'{col_letter}{row}'].value
            if row == 1:
                copied_rows = ws_stare_source.max_row - 1
    
    print(f"    ✓ Skopiowano {copied_rows} wierszy (kolumny A:F)")

# ============================================================
# KROK 7: VLOOKUP do stare_oznaczenie
# ============================================================
if propozycja_col and 'stary' in wb_propozycje.sheetnames:
    print(f"\n[10] Wypełnianie kolumny stare_oznaczenie (VLOOKUP)...")
    ws_stary = wb_propozycje['stary']
    
    vlookup_count = 0
    not_found_count = 0
    
    for row in range(2, ws_propozycje.max_row + 1):
        propozycja = ws_propozycje[f'{propozycja_col}{row}'].value
        if propozycja and not is_cell_empty(propozycja):
            skrot_val = ws_propozycje[f'{skrot_col}{row}'].value
            if skrot_val and not is_cell_empty(skrot_val):
                # VLOOKUP w arkuszu stary (kolumna C zawiera skrot, kolumna F zawiera stare_oznaczenie)
                found = False
                for stary_row in range(2, ws_stary.max_row + 1):
                    if ws_stary[f'C{stary_row}'].value == skrot_val:
                        ws_propozycje[f'{stare_oznaczenie_col}{row}'].value = ws_stary[f'F{stary_row}'].value
                        vlookup_count += 1
                        found = True
                        
                        # Debug: pokaż pierwsze 5 przykładów
                        if vlookup_count <= 5:
                            print(f"    ✓ Wiersz {row}: skrot='{skrot_val}' -> stare='{ws_stary[f'F{stary_row}'].value}'")
                        break
                
                if not found:
                    not_found_count += 1
                    if not_found_count <= 3:
                        print(f"    ✗ Wiersz {row}: Nie znaleziono skrótu '{skrot_val}' w arkuszu stary")
    
    print(f"\n    ✓ Znaleziono {vlookup_count} dopasowań")
    if not_found_count > 0:
        print(f"    ⚠ Nie znaleziono {not_found_count} wpisów w słowniku")

# ============================================================
# ZAPISYWANIE
# ============================================================
output_file = 'propozycje_updated.xlsx'
print(f"\n[11] Zapisywanie pliku '{output_file}'...")
wb_propozycje.save(output_file)

print("\n" + "="*60)
print("✓ UKOŃCZONO POMYŚLNIE")
print("="*60)
print(f"\n📊 PODSUMOWANIE:")
print(f"  ZEFZEV: {zefzev_filled} wierszy")
if propozycja_col:
    print(f"  SPLIT propozycji: {split_count} wierszy")
    print(f"  skrot (zlecenie,rysunek): {skrot_filled} wierszy")
print(f"  nowe_oznaczenie: {nowe_oznaczenie_filled} wierszy")
if 'vlookup_count' in locals():
    print(f"  stare_oznaczenie (VLOOKUP): {vlookup_count} dopasowań")
    if 'not_found_count' in locals() and not_found_count > 0:
        print(f"                            {not_found_count} nie znaleziono")
print(f"\n📄 Plik wyjściowy: {output_file}")
print(f"   Oryginalne kolumny: {original_max_col}")
print(f"   Po przetworzeniu: {ws_propozycje.max_column}")
print(f"   Dodano: {ws_propozycje.max_column - original_max_col} nowych kolumn")
