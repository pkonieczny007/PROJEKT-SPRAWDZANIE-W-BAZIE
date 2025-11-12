from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import shutil
import os
import csv

def is_cell_empty(value):
    """
    Sprawdza czy wartość komórki jest naprawdę pusta.
    """
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == '':
        return True
    return False

# File paths
source_file = 'propozycje_updated.xlsx'
output_file = 'propozycje_updated_z_dopasowaniem.xlsx'
dictionary_export_file = 'slownik_stary_C_F.csv'

# Sprawdź czy plik źródłowy istnieje
if not os.path.exists(source_file):
    print(f"BŁĄD: Nie znaleziono pliku '{source_file}'")
    print("Najpierw uruchom skrypt 'sprawdzanie_rewizji.py'")
    exit()

# Kopiuj plik
print(f"Kopiowanie pliku '{source_file}' -> '{output_file}'...")
shutil.copy(source_file, output_file)

# KLUCZOWE: Wczytaj plik Z data_only=True aby czytać WARTOŚCI zamiast FORMUŁ
print(f"Wczytywanie pliku '{output_file}' z data_only=True (tylko wartości, nie formuły)...")
wb_read = load_workbook(output_file, data_only=True)
ws_read = wb_read.active

# Sprawdź czy arkusz 'stary' istnieje
if 'stary' not in wb_read.sheetnames:
    print("BŁĄD: Nie znaleziono arkusza 'stary'!")
    exit()

ws_stary_read = wb_read['stary']

# KROK 1: Tworzenie słownika z WARTOŚCI (nie formuł) z arkusza 'stary'
print("\n=== KROK 1: Tworzenie słownika z arkusza 'stary' (kolumny C -> F) ===")
print("  ℹ Czytanie WARTOŚCI zamiast formuł (data_only=True)")
stary_dict = {}
empty_keys = 0
empty_values = 0
duplicate_keys = 0

for row in range(2, ws_stary_read.max_row + 1):
    # Czytamy WARTOŚCI (nie formuły) z kolumn C i F
    key_c = ws_stary_read[f'C{row}'].value
    value_f = ws_stary_read[f'F{row}'].value
    
    if not is_cell_empty(key_c):
        # Normalizuj klucz
        key_normalized = str(key_c).strip()
        
        # Sprawdź czy wartość F nie jest pusta
        if is_cell_empty(value_f):
            empty_values += 1
            # Debug: pokaż pierwsze 5 przypadków pustych wartości
            if empty_values <= 5:
                print(f"  ⚠ Wiersz {row}: Klucz '{key_normalized}' ma pustą wartość F")
        
        # Sprawdź duplikaty
        if key_normalized in stary_dict:
            duplicate_keys += 1
        
        stary_dict[key_normalized] = value_f
    else:
        empty_keys += 1

print(f"  ✓ Utworzono słownik z {len(stary_dict)} wpisami")
print(f"  ℹ Pominięto {empty_keys} pustych kluczy (kolumna C)")
print(f"  ⚠ Znaleziono {empty_values} pustych wartości (kolumna F)")
print(f"  ℹ Znaleziono {duplicate_keys} duplikatów kluczy")

# EKSPORT SŁOWNIKA DO CSV
print(f"\n=== EKSPORT SŁOWNIKA DO CSV: {dictionary_export_file} ===")
with open(dictionary_export_file, 'w', newline='', encoding='utf-8') as csvfile:
    writer = csv.writer(csvfile, delimiter=';')
    writer.writerow(['Klucz_C', 'Wartość_F', 'Czy_pusta'])
    
    # Sortujemy słownik alfabetycznie
    for key in sorted(stary_dict.keys()):
        is_empty = "PUSTA" if is_cell_empty(stary_dict[key]) else "OK"
        writer.writerow([key, stary_dict[key], is_empty])

print(f"  ✓ Zapisano {len(stary_dict)} wpisów do pliku '{dictionary_export_file}'")

# Przykładowe wpisy ze słownika
print(f"\n  ℹ Przykładowe wpisy ze słownika (pierwsze 10):")
for i, (k, v) in enumerate(list(stary_dict.items())[:10]):
    status = "✓" if not is_cell_empty(v) else "✗ PUSTA"
    print(f"    [{i+1}] '{k}' -> '{v}' {status}")

# TERAZ wczytaj plik PONOWNIE bez data_only aby móc ZAPISYWAĆ
print(f"\n=== Wczytywanie pliku ponownie do ZAPISU ===")
wb_write = load_workbook(output_file)
ws_write = wb_write.active

# Znajdź kolumny stare_oznaczenie i skrot
print("\nSzukanie kolumn 'stare_oznaczenie' i 'skrot'...")
stare_oznaczenie_col = None
skrot_col = None

for col in range(1, ws_write.max_column + 1):
    col_letter = get_column_letter(col)
    header = ws_write[f'{col_letter}1'].value
    if header == 'stare_oznaczenie':
        stare_oznaczenie_col = col_letter
        print(f"  Znaleziono 'stare_oznaczenie' w kolumnie: {col_letter}")
    elif header == 'skrot':
        skrot_col = col_letter
        print(f"  Znaleziono 'skrot' w kolumnie: {skrot_col}")

if not stare_oznaczenie_col or not skrot_col:
    print("BŁĄD: Nie znaleziono wymaganych kolumn!")
    exit()

# KROK 2: Zbieranie wartości skrot
print(f"\n=== KROK 2: Zbieranie wartości z kolumny 'skrot' ({skrot_col}) ===")
skrot_values = []
for row in range(2, min(ws_write.max_row + 1, 12)):
    skrot_val = ws_write[f'{skrot_col}{row}'].value
    if not is_cell_empty(skrot_val):
        skrot_normalized = str(skrot_val).strip()
        exists = "✓ ISTNIEJE" if skrot_normalized in stary_dict else "✗ BRAK"
        has_value = ""
        if skrot_normalized in stary_dict:
            if is_cell_empty(stary_dict[skrot_normalized]):
                has_value = " (ale wartość F jest PUSTA)"
        skrot_values.append((row, skrot_normalized, exists + has_value))

print(f"  ℹ Przykładowe wartości skrot (pierwsze {len(skrot_values)}):")
for row, val, exists in skrot_values:
    print(f"    Wiersz {row}: '{val}' -> {exists}")

# EKSPORT WARTOŚCI SKROT
skrot_export_file = 'wartosci_skrot_do_sprawdzenia.csv'
print(f"\n=== EKSPORT WARTOŚCI SKROT DO CSV: {skrot_export_file} ===")
with open(skrot_export_file, 'w', newline='', encoding='utf-8') as csvfile:
    writer = csv.writer(csvfile, delimiter=';')
    writer.writerow(['Wiersz', 'Wartość_skrot', 'Istnieje_w_słowniku', 'Wartość_F', 'Status'])
    
    for row in range(2, ws_write.max_row + 1):
        skrot_val = ws_write[f'{skrot_col}{row}'].value
        if not is_cell_empty(skrot_val):
            skrot_normalized = str(skrot_val).strip()
            exists = "TAK" if skrot_normalized in stary_dict else "NIE"
            value_f = stary_dict.get(skrot_normalized, "BRAK")
            status = "OK" if exists == "TAK" and not is_cell_empty(value_f) else "PROBLEM"
            writer.writerow([row, skrot_normalized, exists, value_f, status])

print(f"  ✓ Zapisano wartości skrot do pliku '{skrot_export_file}'")

# KROK 3: Wypełnianie kolumny stare_oznaczenie
print(f"\n=== KROK 3: Dopasowywanie i wypełnianie kolumny 'stare_oznaczenie' ({stare_oznaczenie_col}) ===")
filled_count = 0
not_found_count = 0
empty_value_count = 0
not_found_samples = []

for row in range(2, ws_write.max_row + 1):
    skrot_val = ws_write[f'{skrot_col}{row}'].value
    
    if not is_cell_empty(skrot_val):
        skrot_normalized = str(skrot_val).strip()
        
        if skrot_normalized in stary_dict:
            value_f = stary_dict[skrot_normalized]
            
            # Sprawdź czy wartość F nie jest pusta
            if not is_cell_empty(value_f):
                ws_write[f'{stare_oznaczenie_col}{row}'].value = value_f
                filled_count += 1
                
                if filled_count <= 5:
                    print(f"  ✓ Wiersz {row}: '{skrot_normalized}' -> '{value_f}'")
            else:
                empty_value_count += 1
                if len(not_found_samples) < 10:
                    not_found_samples.append(f"'{skrot_normalized}' (wiersz {row}) - KLUCZ ISTNIEJE ale wartość F PUSTA")
        else:
            not_found_count += 1
            if len(not_found_samples) < 10:
                not_found_samples.append(f"'{skrot_normalized}' (wiersz {row}) - BRAK w słowniku")

print(f"\n=== PODSUMOWANIE ===")
print(f"  ✓ Wypełniono: {filled_count} wierszy")
print(f"  ✗ Nie znaleziono w słowniku: {not_found_count} wierszy")
print(f"  ⚠ Znaleziono klucz ale wartość F pusta: {empty_value_count} wierszy")

if not_found_samples:
    print(f"\n  ℹ Przykłady problemów:")
    for sample in not_found_samples:
        print(f"    - {sample}")

# Zapisz plik
print(f"\n=== ZAPISYWANIE ===")
print(f"Zapisywanie pliku '{output_file}'...")
wb_write.save(output_file)
print(f"✓ Plik został pomyślnie zapisany jako '{output_file}'")
print(f"\n📄 Wygenerowane pliki do analizy:")
print(f"  1. {dictionary_export_file} - słownik z oznaczeniem pustych wartości")
print(f"  2. {skrot_export_file} - analiza dopasowań")
print(f"  3. {output_file} - wynikowy plik Excel")
