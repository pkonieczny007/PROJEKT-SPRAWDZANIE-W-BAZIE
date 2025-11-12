import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import openpyxl.utils.cell
import os

def is_cell_empty(value):
    """
    Sprawdza czy wartość komórki jest naprawdę pusta.
    Zwraca True jeżeli wartość to None, pusty string lub string z samymi spacjami.
    """
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == '':
        return True
    return False

# File paths
propozycje_file = 'propozycje.xlsx'
stare_file = r'\\QNAP-ENERGO\Technologia\BAZA\POBIERANIE Z BAZY\SPRAWDZANIE REWIZJI\SPRAWDZANIE_REWIZJI-STARE.xlsx'

# Wczytaj plik Excel używając pandas
df = pd.read_excel(propozycje_file)

# Automatycznie znajdź kolumny po nazwach
try:
    zeinr_col_idx = df.columns.get_loc("Zeinr")
    zef_col_idx = df.columns.get_loc("ZEF")
    zev_col_idx = df.columns.get_loc("ZEV")
except KeyError as e:
    print(f"Błąd: Nie znaleziono kolumny {e}. Sprawdź nazwy kolumn w pliku.")
    exit()

# Konwertuj indeksy kolumn na litery Excel (dodaj 1, bo Excel jest 1-based)
zeinr_col = get_column_letter(zeinr_col_idx + 1)
zef_col = get_column_letter(zef_col_idx + 1)
zev_col = get_column_letter(zev_col_idx + 1)

print(f"Znaleziono kolumny:")
print(f"  Zeinr: {zeinr_col}")
print(f"  ZEF: {zef_col}")
print(f"  ZEV: {zev_col}")

# Teraz wczytaj plik używając openpyxl do edycji
wb_propozycje = load_workbook(propozycje_file)
ws_propozycje = wb_propozycje.active

# Create ZEFZEV column next to ZEV
zefzev_col = get_column_letter(openpyxl.utils.column_index_from_string(zev_col) + 1)
ws_propozycje[f'{zefzev_col}1'].value = 'ZEFZEV'

# Fill ZEFZEV with ZEF&ZEV - tworzymy JEŻELI przynajmniej jedna wartość jest niepusta
print("\nWypełnianie kolumny ZEFZEV...")
for row in range(2, ws_propozycje.max_row + 1):
    zef = ws_propozycje[f'{zef_col}{row}'].value
    zev = ws_propozycje[f'{zev_col}{row}'].value
    
    # Jeżeli OBE puste - nie tworzymy
    if is_cell_empty(zef) and is_cell_empty(zev):
        ws_propozycje[f'{zefzev_col}{row}'].value = None
    else:
        # Jeżeli przynajmniej jedna niepusta - konkatenujemy
        zef_str = str(zef) if not is_cell_empty(zef) else ''
        zev_str = str(zev) if not is_cell_empty(zev) else ''
        ws_propozycje[f'{zefzev_col}{row}'].value = zef_str + zev_str

# Tworzenie słownika Zeinr -> ZEFZEV (pomijając puste wartości ZEFZEV)
print("\nTworzenie słownika Zeinr -> ZEFZEV...")
zeinr_to_zefzev = {}
for row in range(2, ws_propozycje.max_row + 1):
    zeinr = ws_propozycje[f'{zeinr_col}{row}'].value
    zefzev = ws_propozycje[f'{zefzev_col}{row}'].value
    
    # Dodajemy do słownika tylko gdy ZEFZEV nie jest puste
    if not is_cell_empty(zeinr) and not is_cell_empty(zefzev):
        zeinr_to_zefzev[zeinr] = zefzev

print(f"Utworzono słownik z {len(zeinr_to_zefzev)} wpisami")

# Task 2: Create new column 'nowe_oznaczenie' using ZEFZEV
nowe_oznaczenie_col = get_column_letter(ws_propozycje.max_column + 1)
ws_propozycje[f'{nowe_oznaczenie_col}1'].value = 'nowe_oznaczenie'

# Znajdź kolumnę 'propozycja' automatycznie
try:
    propozycja_col_idx = df.columns.get_loc("propozycja")
    propozycja_col = get_column_letter(propozycja_col_idx + 1)
    print(f"  propozycja: {propozycja_col}")
except KeyError:
    print("Ostrzeżenie: Nie znaleziono kolumny 'propozycja'.")
    propozycja_col = None

# Wypełnianie nowe_oznaczenie na podstawie słownika Zeinr -> ZEFZEV
print("\nWypełnianie kolumny nowe_oznaczenie...")
filled_count = 0
for row in range(2, ws_propozycje.max_row + 1):
    zeinr = ws_propozycje[f'{zeinr_col}{row}'].value
    propozycja = ws_propozycje[f'{propozycja_col}{row}'].value if propozycja_col else True
    
    if not is_cell_empty(propozycja):
        if zeinr in zeinr_to_zefzev:
            ws_propozycje[f'{nowe_oznaczenie_col}{row}'].value = zeinr_to_zefzev[zeinr]
            filled_count += 1

print(f"Wypełniono {filled_count} wierszy w kolumnie nowe_oznaczenie")

# Task 5: Create 'stare_oznaczenie' column next to 'nowe_oznaczenie'
stare_oznaczenie_col = get_column_letter(openpyxl.utils.column_index_from_string(nowe_oznaczenie_col) + 1)
ws_propozycje[f'{stare_oznaczenie_col}1'].value = 'stare_oznaczenie'

# Task 6: Create 'skrot' column next to 'stare_oznaczenie'
skrot_col = get_column_letter(openpyxl.utils.column_index_from_string(stare_oznaczenie_col) + 1)
ws_propozycje[f'{skrot_col}1'].value = 'skrot'

# Task 3: Create new sheet 'stary' and copy columns A:F from stare_file
print("\nKopiowanie danych ze starego pliku...")
wb_stare = load_workbook(stare_file)
ws_stare_source = wb_stare.active

if 'stary' in wb_propozycje.sheetnames:
    wb_propozycje.remove(wb_propozycje['stary'])

ws_stary = wb_propozycje.create_sheet('stary')

# Copy columns A:F
for col in range(1, 7):
    col_letter = get_column_letter(col)
    for row in range(1, ws_stare_source.max_row + 1):
        ws_stary[f'{col_letter}{row}'].value = ws_stare_source[f'{col_letter}{row}'].value

# Task 4: Split 'propozycja' column by '_' i umieść w odległej kolumnie BA
BA_COL_NUM = 53
BG_COL_NUM = 59
ba_col = get_column_letter(BA_COL_NUM)
bg_col = get_column_letter(BG_COL_NUM)

if propozycja_col:
    print(f"\nSplit kolumny propozycja w kolumnie: {ba_col} (i dalej)")
    for row in range(2, ws_propozycje.max_row + 1):
        propozycja = ws_propozycje[f'{propozycja_col}{row}'].value
        if propozycja and not is_cell_empty(propozycja):
            parts = str(propozycja).split('_')
            for i, part in enumerate(parts):
                col_num = BA_COL_NUM + i
                col_letter = get_column_letter(col_num)
                ws_propozycje[f'{col_letter}{row}'].value = part

    # Fill 'skrot' with concatenation: BG & Zeinr
    print(f"\nWypełnianie kolumny skrot ({bg_col} & Zeinr)...")
    skrot_filled = 0
    for row in range(2, ws_propozycje.max_row + 1):
        propozycja = ws_propozycje[f'{propozycja_col}{row}'].value
        if propozycja and not is_cell_empty(propozycja):
            bg = ws_propozycje[f'{bg_col}{row}'].value
            zeinr = ws_propozycje[f'{zeinr_col}{row}'].value
            if bg and zeinr and not is_cell_empty(bg) and not is_cell_empty(zeinr):
                ws_propozycje[f'{skrot_col}{row}'].value = str(bg) + ',' + str(zeinr)
                skrot_filled += 1
    
    print(f"Wypełniono {skrot_filled} wierszy w kolumnie skrot")

# Save the workbook
output_file = 'propozycje_updated.xlsx'
wb_propozycje.save(output_file)
print(f"\n✓ Plik został pomyślnie zapisany jako '{output_file}'")
print("\nUruchom teraz skrypt 'dopasowanie_rewizji.py' aby wypełnić kolumnę stare_oznaczenie")
